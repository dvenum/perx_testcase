#!/usr/bin/env python3
"""
    Dataset generator

    It build one or few .xlsx files of followed format:

    * загружаемый файл представляет собой файл Excel (xls или xlsx)
    * файл состоит из нескольких листов, каждый из которых имеет свой набор колонок или не имеет вообще никаких данных
    * количество колонок, а также их порядок не фиксирован
    * в первой строке указаны заголовки колонок
    * последней колонкой на листе считается та, у которой справа от её заголовка находится пустая ячейка
    * данные в колонке также ограничены пустой ячейкой снизу
    * на одном из листов содержатся колонки с заголовками before и after
    * далее речь идёт о только колонках before и after
        одна из колонок содержит набор L1 из N положительных целых чисел
        другая колонка содержит набор L2 из N+1 положительное целое число
        известно, что набор из L2 состоит из набора L1 и ещё некоего числа X
        порядок следования общих чисел в наборах может различаться


    Remarks:
    * Adapted for Unix urls only
    * There is faker lib (https://github.com/joke2k/faker), thats good at similar case.
      Here I want lot of humans unreadable data.

    Usage:
        $ python3 spawner.py [--dir $DIRECTORY] [$N]

"""

import openpyxl
import random
import argparse
import string

import sys, os

from openpyxl import Workbook


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
BASE_FILENAME = 'dataset_'


COLUMNS_WITH = 26

# restriction of random.choice()
# should never be excessed in normal case
LIMIT_OF_SAFE = 500 

def get_random_name(destdir, base, ext='.xlsx'):
    ''' return new unique filename
        
        ex.: 'dataset_yshwzirk.xlsx'
        looks more fun and natural, I hope
    '''
    POSTFIX_LENGTH = 8

    if not ext.startswith('.'):
        ext = f'.{ext}'

    allowed = string.ascii_lowercase
    for _ in range(LIMIT_OF_SAFE):
        postfix = ''.join(random.choice(allowed) for _ in range(POSTFIX_LENGTH))
        filename = os.path.join(destdir, f'{base}{postfix}{ext}')
        if not os.path.exists(filename):
            return filename

    print('Erorr: Unknown problem in get_random_name()')
    return None


def get_random_meta():
    ''' Return random parameters of sample document
    '''

    return {
        'sheets_count': random.randint(1, 8),
        'sequence_len': random.randint(1000, 18000),
    }


def get_random_str(capitalized=False):
    ''' Random letters of random length
    '''
    MIN_LENGTH = 4
    MAX_LENGTH = 12
    length = random.randint(MIN_LENGTH, MAX_LENGTH)

    allowed = string.ascii_lowercase
    seq = ''.join(random.choice(allowed) for _ in range(length))

    if capitalized:
        seq = seq.capitalize()
    return seq



def get_random_sequence(length):
    ''' make 'before' and 'after', see HEAD for details
    '''
    base_seq = [random.randint(1, sys.maxsize) for _ in range(length)]
    before = base_seq.copy()

    for _ in range(LIMIT_OF_SAFE):
        x = random.randint(1, sys.maxsize)
        if x not in base_seq:
            break

    if x in base_seq:
        return None

    base_seq.append(x)
    random.shuffle(base_seq)
    after = base_seq

    if random.random() >= 0.5:
        return before, after
    else:
        return after, before


def get_random_column_name():
    '''
    '''
    return f'{get_random_str(capitalized=True)} {get_random_str()}'


def fill_sheet(ws):
    ''' Fill random garbage to all sheets
    '''
    is_empty = random.random() < 0.08
    if is_empty:
        return True

    columns_count = random.randint(8, 16)
    
    for col in range(1, columns_count+1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = COLUMNS_WITH
        cname = get_random_column_name()
        cell = ws.cell(row=1, column=col)
        cell.value = cname

        for row in range(2, random.randint(1200, 9800)+1):
            cell = ws.cell(row=row, column=col)
            cell.value = random.randint(1, sys.maxsize)
            cell.number_format = '0'

    return True


def insert_column(ws, col_idx):
    ''' naive columns copier
        instead of native insert_cols(), it save formatting

        It leave format of old column as is
    '''
    col_idx_new = ws.max_column
    row = 1
    ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx_new)].width = \
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width
    while True:
        cell = ws.cell(row=row, column=col_idx)
        cell_new = ws.cell(row=row, column=col_idx_new)
        cell_new.value = cell.value
        cell_new.number_format = cell.number_format
        cell.value = None

        row += 1
        if not ws.cell(row=row, column=col_idx).value:
            break


def fill_column_random_place(ws, column_name, data):
    '''
    '''
    col = random.randint(1, ws.max_column)
    insert_column(ws, col)
    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = COLUMNS_WITH
    ws.cell(row=1, column=col).value = column_name

    row = 2
    for value in data:
        cell = ws.cell(row=row, column=col)
        cell.value = value
        cell.number_format = '0'
        row += 1


def build_datasample():
    ''' Fill .xlsx with random data and write it
    '''
    params = get_random_meta()
    before, after = get_random_sequence(params['sequence_len'])

    wb = Workbook()

    # Prepare random sheets
    ws = wb.active
    ws.title = get_random_str()    # default sheet
    fill_sheet(ws)
    for _ in range(params['sheets_count']-1):
        ws = wb.create_sheet(get_random_str())
        fill_sheet(ws)

    # Add targets columns
    for _ in range(LIMIT_OF_SAFE):
        gems_sheet = wb[random.choice(wb.sheetnames)]
        if gems_sheet.max_column > 1:
            break

    if gems_sheet.max_column > 1:
        fill_column_random_place(gems_sheet, 'before', before)
        fill_column_random_place(gems_sheet, 'after', after)
        return wb

    return None


def main(destdir, count=1):

    for _ in range(count):
        filename = get_random_name(destdir, base=BASE_FILENAME)
        out = build_datasample()
        if not out:
            count += 1
            print('Warning: Incorrect dataset skipped')
            continue
        
        out.save(filename)
    return 0


def create_dir(d):
    ''' create directory recursively
        @return: True, if exists or created, None otherwise
    '''
    if os.path.exists(d):
        return True

    if d.startswith('/'):
        new_dir = d
    else:
        new_dir = os.path.join(BASE_DIR, d)

    try:
        os.makedirs(new_dir)
    except (OSError, PermissionError):
        return False
    else:
        return True


if __name__ == '__main__':

    parser = argparse.ArgumentParser(description='Testcase dataset generator. See head of .py file.')
    parser.add_argument('-d', '--dir', help='Destination dir', type=str, default=BASE_DIR, required=False)
    parser.add_argument('-c', '--count', help='Number of output .xlsx documents', type=int, 
                                                                    default=1, required=False)
    args = parser.parse_args()

    if not create_dir(args.dir):
        print('Error: Incorrect output directory')
        exit(os.EX_OSERR)

    if args.count <= 0:
        print('Error: Positive count allowed only')
        exit(os.EX_USAGE)

    exit(main(args.dir, args.count))



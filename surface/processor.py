#!/usr/bin/env python3
'''
    Work on xls document:

    * необходимо определить число X
    * если набор L1 находился в колонке before, то в результат обработки поместить added: X
    * если набор L1 находился в колонке after, то в результат обработки поместить removed: X
    * время обработки должно быть линейным

    Remark:
        * pylightxl should be faster here, when it doesn't support old .xls format
        (https://pylightxl.readthedocs.io/en/latest/)

'''
import openpyxl
from io import BytesIO

from storage import models


LABEL_BEFORE = 'before'
LABEL_AFTER = 'after'

STATUS_NOTREADY = 'unprocessed'
STATUS_INVALID = 'invalid'
STATUS_ADDED = 'added'
STATUS_REMOVED = 'removed'


class Column(object):
    ''' Describe one specific column on Document

        Now it allow to load set() with all values only,
        but can be extended with data filters and etc. easily
    '''

    def __init__(self, table, col_idx):
        ''' @param table: list of columns ex.: [(A1,A2, ..), (B1, B2, ..)]
        '''

        self.column_index = col_idx
        self.data = set(c.value for c in table[self.column_index-1][1:] if c.value)

    def load_set(self):
        return self.data


class Document(object):
    ''' Describe one xls/xlsx document
    '''

    def __init__(self):
        self.wb = None
        self.path = None
        self.source = None
        self.before = None
        self.after = None
        self._x = None
        self.status = STATUS_NOTREADY
        self.target_table = None    # preloaded list of cells


    def load_by_uuid(self, uuid):
        doc_model = models.DocumentModel.objects.get(id=uuid)
        return self.load_from_model(doc_model)

    def load_from_model(self, doc_model):
        file_obj = doc_model.load()
        self.wb = openpyxl.load_workbook(filename=BytesIO(file_obj.read()))

        self.path = doc_model.filename
        self.source = doc_model.located

        return doc_model

    def load_from_file(self, filename):
        self.wb = openpyxl.load_workbook(filename, read_only=True)
        self.path = filename
        self.source = models.DOCUMENT_LOCATION.LOCAL

    def load_sheet(self, ws):
        ''' this xml document is row oriented and openpyxl can show columns
            only when each row is parsed.
            Here we load sheet entirely by performance reason at cost of memory
        '''
        return list(zip(*ws.rows))

    def discovery(self):
        ''' Find 'before' and 'after' columns
        '''
        for ws in self.wb.worksheets:
            if ws.max_row <= 1: # empty sheet
                continue

            head = [c.value for c in ws[1]]
            if LABEL_BEFORE in head and LABEL_AFTER in head:
                self.target_table = self.load_sheet(ws)
                if not self.target_table:
                    return None

                for col_idx,h in enumerate(head,start=1):
                    if h == LABEL_BEFORE:
                        self.before = Column(self.target_table, col_idx)
                    elif h == LABEL_AFTER:
                        self.after = Column(self.target_table, col_idx)
                break

        return self.before and self.after

    def obtain_difference(self, N1, N2):
        # complexity of difference s-t is O(len(s))
        # https://wiki.python.org/moin/TimeComplexity

        r = N1 - N2
        if len(r) != 1:
            return None
        return r.pop()

    @property
    def x(self):
        ''' Find a targets X, set status
        '''

        if self._x:
            return self._x

        if not all((self.before, self.after)):
            if not self.discovery():
                self.status = STATUS_INVALID
                return None

        n_before = self.before.load_set()
        n_after = self.after.load_set()
        if len(n_before) > len(n_after):
            self.status = STATUS_ADDED
            self._x = self.obtain_difference(n_before, n_after)
        elif len(n_after) > len(n_before): 
            self.status = STATUS_REMOVED
            self._x = self.obtain_difference(n_after, n_before)
        else:
            self.status = STATUS_INVALID
            return None

        if not self._x:
            self.status = STATUS_INVALID
        return self._x


if __name__ == '__main__':
    test_doc = ['tests/testdata/dataset_fjveqerc.xlsx',
                'tests/testdata/dataset_atyoellt.xlsx',
               ]

    for t in test_doc:
        doc = Document()
        doc.load_from_file(t)
        print(t, doc.x, doc.status)


